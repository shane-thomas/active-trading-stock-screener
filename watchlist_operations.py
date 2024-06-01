import os
import constants as c
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from random import randint


def rand_color():
    alpha = "40"  # 50% transparent
    red = format(randint(128, 255), '02X')
    green = format(randint(128, 255), '02X')
    blue = format(randint(128, 255), '02X')
    return f"{alpha}{red}{green}{blue}"


def update_workbook(sheet_name, header_columns):
    workbook = load_workbook(c.ROC_FILE)
    worksheet = workbook[sheet_name]

    if worksheet.max_row == 1:
        for col_num, column_title in enumerate(header_columns, 1):
            worksheet.cell(row=1, column=col_num, value=column_title)

    if sheet_name == '5 DAYS' and worksheet.max_row == 251:
        worksheet.move_range("A52:P251", rows=-50)

    if sheet_name == ('30 DAYS' or '90 DAYS') and worksheet.max_row == 751:
        worksheet.move_range("A52:P751", rows=-50)

    workbook.save(c.ROC_FILE)
    workbook.close()


def color_cells(sheet_name: str, start_row: int):
    argb = rand_color()
    workbook = load_workbook(c.ROC_FILE)
    worksheet = workbook[sheet_name]
    for row in range(start_row+1, start_row + 51):
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row, column=col).fill = PatternFill(
                start_color=argb, end_color=argb, fill_type='solid')

    workbook.save(c.ROC_FILE)
    workbook.close()


def get_symbols_subset(directory: str) -> list[str]:
    sheet = directory.split('/')[1]
    worksheet = pd.read_excel(c.ROC_FILE, sheet_name=sheet)
    symbols = worksheet['SYMBOL'].tolist()
    return sorted(set(symbols))


def calculate_roc(directory: str):
    sheet_name = directory.split('/')[1]

    roc_dict = {}
    dataframes = []

    files_list = [file for file in os.listdir(directory) if file.endswith(".csv")]
    
    current_report = pd.read_csv(os.path.join(directory, files_list[-1]))
    current_report = current_report.query('SERIES == "EQ"').reset_index(drop=True)

    
    os.system('cls')

    for file in files_list:
        file_path = os.path.join(directory, file)
        df = pd.read_csv(file_path)
        df = df.query('SERIES == "EQ"')
        dataframes.append(df[['SYMBOL', 'CLOSE']])

    combined_df = pd.concat(dataframes)

    low_dict = combined_df.groupby('SYMBOL')['CLOSE'].min().to_dict()
        
    os.system('cls')
    for symbol in tqdm(current_report["SYMBOL"], desc=f"Generating report of last {sheet_name}"):
        report_index = current_report.index.get_loc(current_report[current_report['SYMBOL'] == symbol].index[0])
        current_close = current_report.at[report_index, 'CLOSE']
        roc = ((current_close - low_dict[symbol]) / current_close) * 100
        roc_dict[symbol] = roc


    current_report.insert(len(current_report.columns), 'ROC', value=current_report['SYMBOL'].map(roc_dict))
    current_report.insert(len(current_report.columns), 'LC', value=current_report['SYMBOL'].map(low_dict))

    current_report = current_report.sort_values(by='ROC', ascending=True)

    header_columns = current_report.columns.tolist()
    update_workbook(sheet_name, header_columns)

    with pd.ExcelWriter(c.ROC_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        start_row = writer.sheets[sheet_name].max_row
        current_report.head(50).to_excel(
            writer, sheet_name, index=False, startrow=start_row, header=None)

    color_cells(sheet_name, start_row)


def calculate_roc_subset(symbols: list[str], directory: str):
    del_columns = ['TOTTRDVAL', 'TIMESTAMP', 'TOTALTRADES',
                   'ISIN', 'Unnamed: 13', 'LAST', 'OPEN', 'HIGH', 'LOW']

    sheet_name = directory.split('/')[1]

    roc_dict = {}
    dataframes = []

    files_list = [file for file in os.listdir(directory) if file.endswith(".csv")]

    current_report = pd.read_csv(os.path.join(directory, files_list[-1]))
    current_report = current_report.query('SERIES == "EQ"').reset_index(drop=True)

    os.system('cls')

    for file in files_list:
        file_path = os.path.join(directory, file)
        df = pd.read_csv(file_path)
        df = df.query('SERIES == "EQ"')
        dataframes.append(df[['SYMBOL', 'CLOSE']])

    combined_df = pd.concat(dataframes)

    low_dict = combined_df.groupby('SYMBOL')['CLOSE'].min().to_dict()
    
    for symbol in symbols:
        if symbol in df["SYMBOL"].tolist() and symbol in current_report["SYMBOL"].tolist():
            report_index = current_report.index.get_loc(current_report[current_report['SYMBOL'] == symbol].index[0])
            current_close = current_report.at[report_index, 'CLOSE']

            roc = ((current_close - low_dict[symbol]) / current_close) * 100
            roc_dict[symbol] = roc

    d2_report = pd.read_csv(os.path.join(directory, files_list[-2]))
    d3_report = pd.read_csv(os.path.join(directory, files_list[-3]))

    d2 = dict(zip(d2_report['SYMBOL'], d2_report['PREVCLOSE']))
    d3 = dict(zip(d3_report['SYMBOL'], d3_report['PREVCLOSE']))
    current_report.insert(len(current_report.columns),
                          'D2PREVCLOSE', value=current_report['SYMBOL'].map(d2))
    current_report.insert(len(current_report.columns),
                          'D3PREVCLOSE', value=current_report['SYMBOL'].map(d3))

    current_report.insert(len(current_report.columns),'ROC', value=current_report['SYMBOL'].map(roc_dict))

    current_report.insert(len(current_report.columns), 'LC', value=current_report['SYMBOL'].map(low_dict))


    current_report = current_report.sort_values(by='ROC', ascending=True)
    current_report = current_report.query(c.QUERY)

    for col in del_columns:
        current_report.pop(col)

    with pd.ExcelWriter(c.RESULTS_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        current_report.head(50).to_excel(writer, sheet_name, index=False)


if __name__ == "__main__":
    calculate_roc("DATA/5 DAYS", 2)
